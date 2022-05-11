from statistics import mode
from time import gmtime, strftime
from numpy import float64
import xlsxwriter as xlwrite
from EmailExtractor import EMFPlogic, ICClogic
from ExcelExtractor import *
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell
from itertools import chain
import re
import datetime

import json

date = datetime.date.today()


def createWorkbook(excel_dir):

    double = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    header = ['Factory/Site', '', '', '', 'Date', '', '', '', '', '']

    fName = ['CCC4 (CST)', 'CCC2 (CST)', 'CCC6 (CST)',
             'APCC (MYT)', 'ICC (IST)', 'EMFP (CET)', 'BRH1 (BRT)']

    CCC4List = ['DT Kitting&Cell', 'DT Backend', 'SV Kitting&Cell K6',
                'SV Kitting&Cell K7', 'SV Backend', 'Storage line', 'CFS']

    CCC2List = ['DT Kitting&Cell', 'DT Backend', 'SV Kitting&Cell',
                'SV Backend', 'K8 line', 'ARB']

    APCCList = ['Desktop', 'HYBRID 1', 'HYBRID 2', 'Server']

    ICCList = ['Line 1 Frontend', 'Line 2 Frontend', 'Line 3 Frontend',
               'Line 1 Backend', 'Line 2 Backend', 'Line 3 Backend']

    BRHList = ['Notebook', 'Desktop', 'Server', 'AIO']

    CCC6List = ['Line 1', 'Line 2', 'Line 3', 'Line 4', 'Line 5',
                'Line 6', 'Line 7', 'Line 8', 'Line 9', 'Line 10']

    subheader = ['Line', 'Start Time', 'End Time',
                 'UPH', 'Start Time', 'End Time', 'UPH', 'Start Time', 'End Time', 'UPH']

    wb = Workbook()
    ws = wb.active

    # title of worksheet
    ws.title = "Workplans"

    # create legend
    ws['M7'] = 'Legend'
    ws['M7'].font = Font(bold=True)
    ws.merge_cells('M7:N7')
    ws['M7'].border = double
    ws['N7'].border = double

    ws['M8'] = 'First shift'
    ws['M9'] = 'Second shift'
    ws['M10'] = 'Third shift'
    ws['M8'].border = double
    ws['M9'].border = double
    ws['M10'].border = double

    ws['N8'].fill = PatternFill("solid", fgColor="00FFFFCC")
    ws['N9'].fill = PatternFill("solid", fgColor="00FFFF00")
    ws['N10'].fill = PatternFill("solid", fgColor="8DB4E2")
    ws['N8'].border = double
    ws['N9'].border = double
    ws['N10'].border = double

    # title of sheet
    ws.merge_cells(start_column=3, start_row=3, end_column=9, end_row=4)
    ws['C3'] = "Consolidated Factory Workplan"
    ws['C3'].font = Font(b=True, size=18)
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')

    # updated on info
    ws.merge_cells(start_column=3, start_row=5, end_column=5, end_row=5)
    ws['C5'] = "Updated on :"
    ws['C5'].alignment = Alignment(horizontal='left')

    ws.merge_cells(start_column=6, start_row=5, end_column=8, end_row=5)

    # subheader and cell color
    for col in ws.iter_cols(min_col=2, min_row=8, max_col=2, max_row=63):
        for cell in col:
            cell.fill = PatternFill("solid", fgColor="DDDDDD")

    for col in ws.iter_cols(min_col=3, min_row=8, max_col=5, max_row=63):
        for cell in col:
            cell.fill = PatternFill("solid", fgColor="00FFFFCC")

    for col in ws.iter_cols(min_col=6, min_row=8, max_col=8, max_row=63):
        for cell in col:
            cell.fill = PatternFill("solid", fgColor="00FFFF00")

    for col in ws.iter_cols(min_col=9, min_row=8, max_col=11, max_row=63):
        for cell in col:
            cell.fill = PatternFill("solid", fgColor="8DB4E2")

    # add factory lines
        # CCC4 and CCC2
    # merge for night UPH
    ws.merge_cells(start_column=8, start_row=11, end_column=8, end_row=13)

    ws.merge_cells(start_column=8, start_row=18, end_column=8, end_row=19)
    ws.merge_cells(start_column=8, start_row=20, end_column=8, end_row=21)

    # merge for day UPH
    ws.merge_cells(start_column=5, start_row=9, end_column=5, end_row=10)
    ws.merge_cells(start_column=5, start_row=11, end_column=5, end_row=13)

    ws.merge_cells(start_column=5, start_row=18, end_column=5, end_row=19)
    ws.merge_cells(start_column=5, start_row=20, end_column=5, end_row=21)

    for col in ws.iter_cols(min_col=2, min_row=9, max_col=2, max_row=8 + len(CCC4List)):
        i = 0
        for cell in col:
            cell.value = CCC4List[i]
            i += 1

    for col in ws.iter_cols(min_col=2, min_row=18, max_col=2, max_row=17 + len(CCC2List)):
        i = 0
        for cell in col:
            cell.value = CCC2List[i]
            i += 1

        # APCC
    for col in ws.iter_cols(min_col=2, min_row=34, max_col=2, max_row=33 + len(APCCList)):
        i = 0
        for cell in col:
            cell.value = APCCList[i]
            i += 1

        # ICC
    for col in ws.iter_cols(min_col=2, min_row=42, max_col=2, max_row=41 + len(ICCList)):
        i = 0
        for cell in col:
            cell.value = ICCList[i]
            i += 1

        # BRH
    for col in ws.iter_cols(min_col=2, min_row=58, max_col=2, max_row=57 + len(BRHList)):
        i = 0
        for cell in col:
            cell.value = BRHList[i]
            i += 1

    # create tables
    rows = 7
    findex = 0

    for row in ws.iter_rows(min_row=7, min_col=2, max_row=63, max_col=11):
        for cell in row:
            cell.border = double

        if rows == 7 or rows == 16 or rows == 24 or rows == 32 or rows == 40 or rows == 48 or rows == 56:
            i = 0

            for cell in row:
                cell.value = header[i]
                # cell.border = double
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="00FFCC99")
                i += 1

            # factory name
            ws.cell(row=rows, column=4, value=fName[findex])
            findex += 1

        elif rows == 8 or rows == 17 or rows == 25 or rows == 33 or rows == 41 or rows == 49 or rows == 57:
            i = 0
            for cell in row:
                cell.value = subheader[i]
                cell.font = Font(bold=True)
                i += 1

        rows += 1

    # merge header
    header = 7
    max_row = 63

    for x in range(max_row):
        if header == 7 or header == 16 or header == 24 or header == 32 or header == 40 or header == 48 or header == 56:
            ws.merge_cells(start_row=header, start_column=2,
                           end_row=header, end_column=3)
            ws.merge_cells(start_row=header, start_column=4,
                           end_row=header, end_column=5)
            ws.merge_cells(start_row=header, start_column=6,
                           end_row=header, end_column=9)
            ws.merge_cells(start_row=header, start_column=10,
                           end_row=header, end_column=11)

        header += 1

    # save xl to explorer
    wb.save(excel_dir)


def CCC4DataInsert(factDf, excel_dir):
    wb = load_workbook(excel_dir)
    ws = wb.active

    df, fname, fdate, fshift, isNight = factDf

    # start shift
    if fname == 'CCC4' and fshift == 'start':

        if isNight:
            # gather data
            ws['J7'] = fdate
            K6_df = df[df['Line'].str.contains("Kitting&Cell K6")]
            K6_start = K6_df.loc[K6_df.first_valid_index(), 'Start Time']

            K7_df = df[df['Line'].str.contains("Kitting&Cell K7")]
            K7_start = K7_df.loc[K7_df.first_valid_index(), 'Start Time']

            SVbackend_df = df[df['Line'].str.contains("SV Backend")]
            SVbackend_start = SVbackend_df.loc[SVbackend_df.first_valid_index(
            ), 'Start Time']

            uph = df.loc[0, 'UPH']

            start_time = [K6_start, K7_start, SVbackend_start, uph]

            # insert data
            ws['H11'] = start_time[3]
            ws['H11'].alignment = Alignment(
                horizontal='center', vertical='center')

            for col in ws.iter_cols(min_col=6, min_row=11, max_col=6, max_row=13):
                i = 0
                for cell in col:
                    cell.value = start_time[i]
                    cell.alignment = Alignment(horizontal='center')

                    i += 1

        else:
            # else is day
            ws['J7'] = fdate

            DTFront_df = df[df['Line'].str.contains("DT Kitting&Cell")]
            DTFront_start = DTFront_df.loc[DTFront_df.first_valid_index(
            ), 'Start Time']

            DTBack_df = df[df['Line'].str.contains("DT Backend")]
            DTBack_start = DTBack_df.loc[DTBack_df.first_valid_index(
            ), 'Start Time']

            K6_df = df[df['Line'].str.contains("Kitting&Cell K6")]
            K6_start = K6_df.loc[K6_df.first_valid_index(), 'Start Time']

            K7_df = df[df['Line'].str.contains("Kitting&Cell K7")]
            K7_start = K7_df.loc[K7_df.first_valid_index(), 'Start Time']

            SVbackend_df = df[df['Line'].str.contains("SV Backend")]
            SVbackend_start = SVbackend_df.loc[SVbackend_df.first_valid_index(
            ), 'Start Time']

            storage_df = df[df['Line'].str.contains("Storage line")]
            storage_start = storage_df.loc[storage_df.first_valid_index(
            ), 'Start Time']

            CFS_df = df[df['Line'].str.contains("CFS")]
            CFS_start = CFS_df.loc[CFS_df.first_valid_index(), 'Start Time']

            uph = [df.loc[0, 'UPH'], df.loc[2, 'UPH']]

            # print(df)

            # store data in array

            start_time = [DTFront_start, DTBack_start, K6_start,
                          K7_start, SVbackend_start, storage_start, CFS_start]

            # insert UPH
            ws['E9'] = uph[0]
            ws['E11'] = uph[1]

            ws['E9'].alignment = Alignment(
                horizontal='center', vertical='center')

            ws['E11'].alignment = Alignment(
                horizontal='center', vertical='center')

            # insert shift time

            for col in ws.iter_cols(min_col=3, min_row=9, max_col=3, max_row=15):
                i = 0
                for cell in col:
                    cell.value = start_time[i]
                    cell.alignment = Alignment(horizontal='center')

                    i += 1

    # Night end shift
    elif fname == 'CCC4' and fshift == 'end':
        if isNight:
            # gather data
            K6_df = df[df['Line'].str.contains("Kitting&Cell K6")]
            K6_end = K6_df.loc[K6_df.first_valid_index(), 'End shift']

            K7_df = df[df['Line'].str.contains("Kitting&Cell K7")]
            K7_end = K7_df.loc[K7_df.first_valid_index(), 'End shift']

            SVbackend_df = df[df['Line'].str.contains("SV Backend")]
            SVbackend_end = SVbackend_df.loc[SVbackend_df.first_valid_index(
            ), 'End shift']

            end_time = [K6_end, K7_end, SVbackend_end]

            # insert data
            for col in ws.iter_cols(min_col=7, min_row=11, max_col=7, max_row=13):
                i = 0
                for cell in col:
                    cell.value = end_time[i]
                    cell.alignment = Alignment(horizontal='center')

                    i += 1

        # else is day
        else:
            # else is day
            ws['J7'] = fdate

            DTFront_df = df[df['Line'].str.contains("DT Kitting&Cell")]
            DTFront_end = DTFront_df.loc[DTFront_df.first_valid_index(
            ), 'End shift']

            DTBack_df = df[df['Line'].str.contains("DT Backend")]
            DTBack_end = DTBack_df.loc[DTBack_df.first_valid_index(
            ), 'End shift']

            K6_df = df[df['Line'].str.contains("Kitting&Cell K6")]
            K6_end = K6_df.loc[K6_df.first_valid_index(), 'End shift']

            K7_df = df[df['Line'].str.contains("Kitting&Cell K7")]
            K7_end = K7_df.loc[K7_df.first_valid_index(), 'End shift']

            SVbackend_df = df[df['Line'].str.contains("SV Backend")]
            SVbackend_end = SVbackend_df.loc[SVbackend_df.first_valid_index(
            ), 'End shift']

            storage_df = df[df['Line'].str.contains("Storage line")]
            storage_end = storage_df.loc[storage_df.first_valid_index(
            ), 'End shift']

            CFS_df = df[df['Line'].str.contains("CFS")]
            CFS_end = CFS_df.loc[CFS_df.first_valid_index(), 'End shift']

            # print(df)

            # store data in array

            end_time = [DTFront_end, DTBack_end, K6_end,
                        K7_end, SVbackend_end, storage_end, CFS_end]

            # insert shift time

            for col in ws.iter_cols(min_col=4, min_row=9, max_col=4, max_row=15):
                i = 0
                for cell in col:
                    cell.value = end_time[i]
                    cell.alignment = Alignment(horizontal='center')

                    i += 1

    wb.save(excel_dir)


def CCC2DataInsert(factDf, excel_dir):
    wb = load_workbook(excel_dir)
    ws = wb.active

    df, fname, fdate, fshift, isNight = factDf

    # start shift
    if fname == 'CCC2' and fshift == 'start':

        if isNight:
            # gather data
            ws['J16'] = fdate
            DTFront_df = df[df['Line'].str.contains("DT Kitting&Cell")]
            DTFront_start = DTFront_df.loc[DTFront_df.first_valid_index(
            ), 'Start Time']

            DTBack_df = df[df['Line'].str.contains("DT Backend")]
            DTBack_start = DTBack_df.loc[DTBack_df.first_valid_index(
            ), 'Start Time']

            SVFront_df = df[df['Line'].str.contains("SV Kitting&Cell")]
            SVFront_start = SVFront_df.loc[SVFront_df.first_valid_index(
            ), 'Start Time']

            SVBack_df = df[df['Line'].str.contains("SV Backend")]
            SVbackend_start = SVBack_df.loc[SVBack_df.first_valid_index(
            ), 'Start Time']

            uph = [df.loc[0, 'UPH'], df.loc[2, 'UPH']]

            start_time = [DTFront_start, DTBack_start, SVFront_start,
                          DTBack_start]

            # insert data
            ws['H18'] = uph[0]
            ws['H20'] = uph[1]

            ws['H18'].alignment = Alignment(
                horizontal='center', vertical='center')

            ws['H20'].alignment = Alignment(
                horizontal='center', vertical='center')

            for col in ws.iter_cols(min_col=6, min_row=18, max_col=6, max_row=(17 + len(start_time))):
                i = 0
                for cell in col:
                    cell.value = start_time[i]
                    cell.alignment = Alignment(horizontal='center')

                    i += 1

        else:
            # else is day
            ws['J16'] = fdate

            DTFront_df = df[df['Line'].str.contains("DT Kitting")]
            DTFront_start = DTFront_df.loc[DTFront_df.first_valid_index(
            ), 'Start Time']

            DTBack_df = df[df['Line'].str.contains("DT Backend")]
            DTBack_start = DTBack_df.loc[DTBack_df.first_valid_index(
            ), 'Start Time']

            SVFront_df = df[df['Line'].str.contains("SV Kitting")]
            SVFront_start = SVFront_df.loc[SVFront_df.first_valid_index(
            ), 'Start Time']

            SVBack_df = df[df['Line'].str.contains("SV Backend")]
            SVbackend_start = SVBack_df.loc[SVBack_df.first_valid_index(
            ), 'Start Time']

            # OT
            K8_start = ''

            if not df[df['Line'].str.contains("K8")].empty:
                K8_df = df[df['Line'].str.contains("K8")]
                K8_start = K8_df.loc[K8_df.first_valid_index(
                ), 'Start Time']

            ARB_df = df[df['Line'].str.contains("ARB")]
            ARB_start = ARB_df.loc[ARB_df.first_valid_index(
            ), 'Start Time']

            uph = [df.loc[0, 'UPH'], df.loc[2, 'UPH']]

            # print(df)

            # store data in array

            start_time = [DTFront_start, DTBack_start,
                          SVFront_start, SVbackend_start, K8_start, ARB_start]

            # insert UPH
            ws['E18'] = uph[0]
            ws['E20'] = uph[1]

            ws['E18'].alignment = Alignment(
                horizontal='center', vertical='center')

            ws['E20'].alignment = Alignment(
                horizontal='center', vertical='center')

            # insert shift time

            for col in ws.iter_cols(min_col=3, min_row=18, max_col=3, max_row=(17+len(start_time))):
                i = 0
                for cell in col:
                    cell.value = start_time[i]
                    cell.alignment = Alignment(horizontal='center')

                    i += 1

    # Night end shift
    elif fname == 'CCC2' and fshift == 'end':
        if isNight:
            # gather data
            DTFront_df = df[df['Line'].str.contains("DT Kitting&Cell")]
            DTFront_end = DTFront_df.loc[DTFront_df.first_valid_index(
            ), 'End shift']

            DTBack_df = df[df['Line'].str.contains("DT Backend")]
            DTBack_end = DTBack_df.loc[DTBack_df.first_valid_index(
            ), 'End shift']

            SVFront_df = df[df['Line'].str.contains("SV Kitting&Cell")]
            SVFront_end = SVFront_df.loc[SVFront_df.first_valid_index(
            ), 'End shift']

            SVbackend_df = df[df['Line'].str.contains("SV Backend")]
            SVbackend_end = SVbackend_df.loc[SVbackend_df.first_valid_index(
            ), 'End shift']

            # OT
            K8_end = ''

            if not df[df['Line'].str.contains("K8")].empty:
                K8_df = df[df['Line'].str.contains("K8")]
                K8_end = K8_df.loc[K8_df.first_valid_index(
                ), 'End shift']

            end_time = [DTFront_end, DTBack_end,
                        SVFront_end, SVbackend_end, K8_end]

            # insert data
            for col in ws.iter_cols(min_col=7, min_row=18, max_col=7, max_row=(17 + len(end_time))):
                i = 0
                for cell in col:
                    cell.value = end_time[i]
                    cell.alignment = Alignment(horizontal='center')

                    i += 1

        # else is day
        else:
            ws['J16'] = fdate

            DTFront_df = df[df['Line'].str.contains("DT Kitting")]
            DTFront_end = DTFront_df.loc[DTFront_df.first_valid_index(
            ), 'End shift']

            DTBack_df = df[df['Line'].str.contains("DT Backend")]
            DTBack_end = DTBack_df.loc[DTBack_df.first_valid_index(
            ), 'End shift']

            SVFront_df = df[df['Line'].str.contains("SV Kitting")]
            SVFront_end = SVFront_df.loc[SVFront_df.first_valid_index(
            ), 'End shift']

            SVBack_df = df[df['Line'].str.contains("SV Backend")]
            SVbackend_end = SVBack_df.loc[SVBack_df.first_valid_index(
            ), 'End shift']

            # OT
            K8_start = ''

            if not df[df['Line'].str.contains("K8")].empty:
                K8_df = df[df['Line'].str.contains("K8")]
                K8_end = K8_df.loc[K8_df.first_valid_index(
                ), 'End shift']

            ARB_df = df[df['Line'].str.contains("ARB")]
            ARB_end = ARB_df.loc[ARB_df.first_valid_index(
            ), 'End shift']

            # store data in array

            end_time = [DTFront_end, DTBack_end,
                        SVFront_end, SVbackend_end, K8_start, ARB_end]

            # insert shift time

            for col in ws.iter_cols(min_col=4, min_row=18, max_col=4, max_row=(17+len(end_time))):
                i = 0
                for cell in col:
                    cell.value = end_time[i]
                    cell.alignment = Alignment(horizontal='center')

                    i += 1

    wb.save(excel_dir)


def APCCDataInsert(df, excel_dir):
    first_result = []
    second_result = []
    first_shift, second_shift, date = df

    wb = load_workbook(excel_dir)
    ws = wb.active

    ws['J32'] = date

    for i in first_shift:
        first_result.append(i.split(' - '))

    for i in second_shift:
        second_result.append(i.split(' - '))

    first_shift_list = list(chain.from_iterable(zip(*first_result)))
    second_shift_list = list(chain.from_iterable(zip(*second_result)))

    i = 0
    for col in ws.iter_cols(min_col=3, max_col=4, min_row=34, max_row=37):
        for cell in col:
            cell.value = list(first_shift_list)[i]
            i += 1

    i = 0
    for col in ws.iter_cols(min_col=6, max_col=7, min_row=34, max_row=37):
        for cell in col:
            cell.value = list(second_shift_list)[i]
            i += 1

    ws['F5'] = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    ws['F5'].alignment = Alignment(horizontal='left')
    ws['I5'] = strftime("%z", gmtime())

    wb.save(excel_dir)


def ICCDataInsert(df, excel_dir):
    # frontend and backend
    front_df, back_df = df

    front_first_shift = []
    front_second_shift = []
    back_first_shift = []
    back_second_shift = []

    wb = load_workbook(excel_dir)
    ws = wb.active

    # if df has two columns put the second column in the second shift col in the excel
    # while the first one put in the first column

    # FRONTEND
    if {'second_shift'}.issubset(front_df.columns):
        # print(front_df['second_shift'][0].split('–'))

        for i, row in front_df.iterrows():
            # front_result.append(i.split('–'))
            # print(row['first_shift'].split('-'))

            front_first_shift.append(re.split('\-|\–', row['first_shift']))
            front_second_shift.append(re.split('\-|\–', row['second_shift']))

            front_first_list = list(
                chain.from_iterable(zip(*front_first_shift)))
            front_second_list = list(
                chain.from_iterable(zip(*front_second_shift)))

        #print(re.split('\-|\–', row['first_shift']))

        i = 0
        for col in ws.iter_cols(min_col=3, max_col=4, min_row=42, max_row=44):
            for cell in col:
                cell.value = list(front_first_list)[i]
                i += 1

        i = 0
        for col in ws.iter_cols(min_col=6, max_col=7, min_row=42, max_row=44):
            for cell in col:
                cell.value = list(front_second_list)[i]
                i += 1

    else:
        for i, row in front_df.iterrows():

            front_first_shift.append(re.split('\-|\–', row['first_shift']))

            front_first_list = list(
                chain.from_iterable(zip(*front_first_shift)))

        i = 0
        for col in ws.iter_cols(min_col=3, max_col=4, min_row=42, max_row=44):
            for cell in col:
                cell.value = list(front_first_list)[i]
                i += 1

    # BACKEND
    if {'second_shift'}.issubset(back_df.columns):

        for i, row in front_df.iterrows():

            back_first_shift.append(re.split('\-|\–', row['first_shift']))
            back_second_shift.append(re.split('\-|\–', row['second_shift']))

            back_first_list = list(
                chain.from_iterable(zip(*front_first_shift)))
            back_second_list = list(
                chain.from_iterable(zip(*front_second_shift)))

            #print(re.split('\-|\–', row['first_shift']))

        i = 0
        for col in ws.iter_cols(min_col=3, max_col=4, min_row=42, max_row=44):
            for cell in col:
                cell.value = list(back_first_list)[i]
                i += 1

        i = 0
        for col in ws.iter_cols(min_col=6, max_col=7, min_row=45, max_row=47):
            for cell in col:
                cell.value = list(back_second_list)[i]
                i += 1

    else:
        for i, row in back_df.iterrows():

            back_first_shift.append(re.split('\-|\–|a', row['first_shift']))

            back_first_list = list(
                chain.from_iterable(zip(*back_first_shift)))

            for i, n in enumerate(back_first_list):
                if n == 'n':
                    back_first_list[i] = ''

        i = 0
        for col in ws.iter_cols(min_col=3, max_col=4, min_row=45, max_row=47):
            for cell in col:
                cell.value = list(back_first_list)[i]
                i += 1

        ws['J40'] = date.strftime('%d-%b')

    wb.save(excel_dir)


def CCC6DataInsert(excel_dir):
    wb = load_workbook(excel_dir)
    ws = wb.active

    # insert from config file
    # CCC6

    config_path = r"sources\factory_config.json"
    # read config file
    with open(config_path) as config_file:
        config = json.load(config_file)
        config = config['CCC6']

    LINE = config['line']
    DATE = config['date']

    START_SHIFT1 = config['start_time1']
    END_SHIFT1 = config['end_time1']
    UPH1 = config['UPH1']

    START_SHIFT2 = config['start_time2']
    END_SHIFT2 = config['end_time2']
    UPH2 = config['UPH2']

    START_SHIFT3 = config['start_time3']
    END_SHIFT3 = config['end_time3']
    UPH3 = config['UPH3']

    ws['B26'] = LINE
    ws['C26'] = START_SHIFT1
    ws['D26'] = END_SHIFT1
    ws['E26'] = UPH1
    ws['F26'] = START_SHIFT2
    ws['G26'] = END_SHIFT2
    ws['H26'] = UPH2
    ws['I26'] = START_SHIFT3
    ws['J26'] = END_SHIFT3
    ws['K26'] = UPH3

    if DATE == "":
        ws['J24'] = date.strftime('%d-%b')
    else:
        ws['J24'] = DATE

    #print(LINE, START_SHIFT1)

    wb.save(excel_dir)


def BRH1DataInsert(df, excel_dir):
    print("Inserting data for BRH")
    wb = load_workbook(excel_dir)
    ws = wb.active

    first_hrs, second_hrs, first_UPH, second_UPH = df

    config_path = r"sources\factory_config.json"
    # read config file
    with open(config_path) as config_file:
        config = json.load(config_file)
        config = config['BRH1']

    for col in ws.iter_cols(min_col=3, min_row=58, max_col=3, max_row=61):
        for cell in col:
            cell.value = config['first_shift']
            cell.number_format = numbers.FORMAT_DATE_TIMEDELTA

    celindex = 58
    for i in first_hrs:
        ws['D%d' % celindex].value = ('=C58 + TIME(%f,0,0)' % i)
        ws['D%d' % celindex].number_format = numbers.FORMAT_DATE_TIME6

        # start second shift
        ws['F%d' % celindex].value = ws['D%d' % celindex].value
        ws['F%d' % celindex].number_format = numbers.FORMAT_DATE_TIME6
        celindex += 1

    celindex = 58
    for i in second_hrs:
        ws['G%d' % celindex].value = ('=F%d + TIME(%f,0,0)' % (celindex, i))
        ws['G%d' % celindex].number_format = numbers.FORMAT_DATE_TIME6

        celindex += 1
    # UPH insertion
    celindex = 58
    for i in first_UPH:
        ws['E%d' % celindex] = i
        celindex += 1

    celindex = 58
    for i in second_UPH:
        ws['H%d' % celindex] = i
        celindex += 1

    ws['J56'] = date.strftime('%d-%b')

    wb.save(excel_dir)


def EMFPDataInsert(df, excel_dir):
    wb = load_workbook(excel_dir)
    ws = wb.active

    #df = EMFPlogic()

    # print(df)
    #print('6:00' in df.to_string())

    # line name
    ws['B50'] = 'EMFP'

    # date
    ws["J48"] = date.strftime('%d-%b')

    if '6:00' in df.to_string():
        ws['C50'] = '6:00'
        ws['D50'] = '14:00'
        ws['C50'].number_format = numbers.FORMAT_DATE_TIME6
        ws['D50'].number_format = numbers.FORMAT_DATE_TIME6
    else:
        ws['C50'] = 'N/A'
        ws['D50'] = 'N/A'

    if '14:00' in df.to_string():
        ws['F50'] = '14:00'
        ws['G50'] = '22:00'
        ws['F50'].number_format = numbers.FORMAT_DATE_TIME6
        ws['G50'].number_format = numbers.FORMAT_DATE_TIME6
    else:
        ws['F50'] = 'N/A'
        ws['G50'] = 'N/A'

    wb.save(excel_dir)


# EMFPDataInsert()
# CCC6DataInsert()
