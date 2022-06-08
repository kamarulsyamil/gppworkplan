from extractor.EmailExtractor import APCClogic, BRHlogic, EMFPlogic, ICClogic, getTableEmail
from extractor.ExcelExtractor_copy import CCC2Night, CCC4Day, CCC2Day, CCC4Night
import creator.ExcelCreator as ExcelCreator
import os.path
import re
import pandas as pd
import os
import json
from openpyxl import load_workbook
import datetime
from openpyxl.styles import Alignment
from time import gmtime, strftime
import streamlit as st


def main():

    config_path = r"app\configuration\tool_config.json"

    # read config file
    with open(config_path) as config_file:
        config = json.load(config_file)
        file_dir = config['file_dir']
        sharepoint = config["sharepoint"]

    #f = "C:\\Users\\Yusuf\\Documents\\My Project\\Factory Work Plan\\Production Line Arrangement of 2022.xlsx"
    #f = r"C:\Users\Yusuf\Documents\My Project\Factory Work Plan\ExcelExtractor\sources\Production Line Arrangement of 2022.xlsx"

    # r'\\w1039fnf93.dhcp.apac.dell.com\PlannerDoc\SHIFT ARRANGEMENT\Production Line Arrangement of 2022.xlsx' #file_dir['CCC2/4']

    f = r'sources\Production Line Arrangement of 2022.xlsx'

    # create workbook named Consolidated Factory Workplan
    if not os.path.exists(file_dir['main_excel']):
        print("Creating workbook...")
        ExcelCreator.createWorkbook(file_dir['main_excel'])

    print("Data processing started...")

    # try:
    print("inserting data for CCC4...")

    ExcelCreator.CCC4DataInsert(CCC4Day(), file_dir['main_excel'])
    ExcelCreator.CCC4DataInsert(CCC4Night(), file_dir['main_excel'])

    print("inserting data for CCC2...")

    ExcelCreator.CCC2DataInsert(CCC2Day(), file_dir['main_excel'])
    ExcelCreator.CCC2DataInsert(CCC2Night(), file_dir['main_excel'])

    # except PermissionError as e:
    #     print("Theres a problem while saving the excel file. Close file if its active.")
    #     print(str(e))

    # except Exception as e:
    #     print("Theres a problem while inserting CCC2/4 data")
    #     print("Error: ", str(e))

    # CCC6 data insertion
    try:
        ExcelCreator.CCC6DataInsert(file_dir['main_excel'])
    except:
        print("Error while reading config file")

    # FROM EMAIL
    # ---------------------------------------------------

    # APCC data insertion

    # drop duplicates and NA

    # ICC BRH data insertion

    try:
        print("Inserting data for APCC workplan")

        ICC_BRH, APCC = getTableEmail()

        for i in APCC:
            factName = re.search('(Work Plan)', i.Body)

            if factName != None:
                if factName.group(0) == 'Work Plan':
                    data1 = pd.read_html(i.HTMLBody)[0].dropna(
                        axis=1, how='all', thresh=3)

                    data1.columns = ['Date', 'Line', 'Frontend', 'Backend']

        if not data1[data1['Date'].astype(str).str.contains("Date")].empty:
            data3 = data1.drop(data1.index[range(5)])
            # print(data3.reset_index(drop=True))
            df = data3.reset_index(drop=True)

            # insert data for APCC
            ExcelCreator.APCCDataInsert(
                APCClogic(df), file_dir['main_excel'])
            print("Successfully inserted data for APCC.")

    except Exception as e:
        print("Error :" + str(e))

    try:
        print("Inserting data for ICC,APCC and BRH workplans")

        for i in ICC_BRH:

            # perlu diubah Tom shift = ICC, Work Plan = APCC,
            factName = re.search(
                '(Tom Shift|Commit Produção)', i.Subject)

            # EMFP OT
            Ot_EMFP = re.search('(EMFP Overtime)', i.Subject)

            # dataframe of table from email
            #data = pd.read_html(i.HTMLBody)

            # print(i.Body)

            #data1 = data1.drop_duplicates(subset='0')

            if factName != None:

                # ICC
                if factName.group(0) == 'Tom Shift':
                    print("Found ICC workplan")

                    # change header of datarframe
                    new_header = pd.read_html(i.HTMLBody)[-1].iloc[0]
                    df = pd.read_html(i.HTMLBody)[-1][1:]
                    df.columns = new_header

                    # date of plan
                    ICC_date = re.findall(
                        r"[0-3][0-9]-[A-Z][a-z][a-z]", i.Body)
                    # print(ICC_date[-1])

                    # insert data for ICC
                    ExcelCreator.ICCDataInsert(
                        ICClogic(df, factName), file_dir['main_excel'], ICC_date[-1])
                    print("Successfully inserted data for ICC.")

                # BRH
                elif factName.group(0) == 'Commit Produção':
                    print("Found BRH workplan")
                    new_header = pd.read_html(i.HTMLBody)[0].iloc[0]
                    df = pd.read_html(i.HTMLBody)[0][1:]
                    df.columns = new_header

                    df1 = df.drop(['LINE', 'CAP', 'Config'], axis=1)

                    new_header = ['LOB', 'HRS1', 'UPH1', 'HRS2', 'UPH2']
                    df1.columns = new_header

                    df1 = df1.fillna(0)

                    ExcelCreator.BRH1DataInsert(
                        BRHlogic(df1, factName), file_dir['main_excel'])
                    print("Successfully inserted data for BRH.")

            elif Ot_EMFP != None:
                print("EMFP OT email found")
                ExcelCreator.OTDataInsert(
                    'EMFP', i.Body, file_dir['main_excel'])

        print("Done!")

    except PermissionError as e:
        print("Theres a problem while saving the excel file. Close file if its active.")
        print(str(e))

    except Exception as e:
        print("Error while processing data from e-mail/No workplan found")
        print(str(e))

    # EMFP insertion
    try:
        print("Inserting data for EMFP...")
        ExcelCreator.EMFPDataInsert(
            EMFPlogic(file_dir["EMFP"]), file_dir['main_excel'])
    except Exception as e:
        print("failed to insert EMFP data")
        print(str(e))

    # Update time
    wb = load_workbook(file_dir['main_excel'])
    ws = wb.active

    ws['F5'] = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    ws['F5'].alignment = Alignment(horizontal='left')
    ws['I5'] = strftime("%z", gmtime())

    wb.save(file_dir['main_excel'])
    print("Successfully inserted data for EMFP.")


if __name__ == "__main__":
    main()
