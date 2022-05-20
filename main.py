from EmailExtractor import APCClogic, BRHlogic, EMFPlogic, ICClogic, getTableEmail
from ExcelExtractor_copy import CCC2Night, CCC4Day, CCC2Day, CCC4Night
import ExcelCreator
import os.path
import re
import pandas as pd
import os
import win32com.client as win32
import json
from openpyxl import load_workbook
import datetime
from openpyxl.styles import Alignment
from time import gmtime, strftime


def main():

    config_path = r"tool_config.json"

    # read config file
    with open(config_path) as config_file:
        config = json.load(config_file)
        file_dir = config['file_dir']
        sharepoint = config["sharepoint"]

    #f = "C:\\Users\\Yusuf\\Documents\\My Project\\Factory Work Plan\\Production Line Arrangement of 2022.xlsx"
    #f = r"C:\Users\Yusuf\Documents\My Project\Factory Work Plan\ExcelExtractor\sources\Production Line Arrangement of 2022.xlsx"

    # r'\\w1039fnf93.dhcp.apac.dell.com\PlannerDoc\SHIFT ARRANGEMENT\Production Line Arrangement of 2022.xlsx' #file_dir['CCC2/4']

    f = r'sources\Production Line Arrangement of 2022.xlsx'

    # create workbook named Consolidated Factory Workplan
    if not os.path.exists(file_dir['main_excel']):
        print("Creating workbook...")
        ExcelCreator.createWorkbook(file_dir['main_excel'])

    print("Data processing started...")

    try:
        print("inserting data for CCC4...")

        ExcelCreator.CCC4DataInsert(CCC4Day(), file_dir['main_excel'])
        ExcelCreator.CCC2DataInsert(CCC2Day(), file_dir['main_excel'])

        ExcelCreator.CCC2DataInsert(CCC2Night(), file_dir['main_excel'])
        ExcelCreator.CCC4DataInsert(CCC4Night(), file_dir['main_excel'])

    except PermissionError as e:
        print("Theres a problem while saving the excel file. Close file if its active.")
        print(str(e))

    except Exception as e:
        print("Theres a problem while inserting CCC2 data")
        print("Error: ", str(e))

    # CCC6 data insertion
    try:
        ExcelCreator.CCC6DataInsert(file_dir['main_excel'])
    except:
        print("Error while reading config file")

    # FROM EMAIL
    # ---------------------------------------------------

    try:
        print("Inserting data for ICC and APCC workplans")

        for i in getTableEmail():

            factName = re.search('(ICC|APCC|BRH)', i.Body)

            # EMFP OT
            Ot_EMFP = re.search('(EMFP Overtime)', i.Subject)

            # dataframe of table from email
            #data = pd.read_html(i.HTMLBody)

            # print(i.Body)

            #data2 = data1.drop_duplicates(subset='0')

            if factName != None:
                if factName.group(0) == 'APCC':
                    # drop duplicates and NA
                    data1 = pd.read_html(i.HTMLBody)[0].dropna(
                        axis=1, how='all', thresh=3)

                    data1.columns = ['Date', 'Line', 'Frontend', 'Backend']

                    data2 = data1  # .drop_duplicates()
                    # print(data2)

                    if not data2[data2['Date'].astype(str).str.contains("Date")].empty:
                        data3 = data2.drop(data2.index[range(5)])
                        # print(data3.reset_index(drop=True))
                        df = data3.reset_index(drop=True)

                        # insert data for APCC
                        ExcelCreator.APCCDataInsert(
                            APCClogic(df, factName), file_dir['main_excel'])

                elif factName.group(0) == 'ICC':
                    # change header of datarframe
                    new_header = pd.read_html(i.HTMLBody)[3].iloc[0]
                    df = pd.read_html(i.HTMLBody)[3][1:]
                    df.columns = new_header

                    # insert data for ICC
                    ExcelCreator.ICCDataInsert(
                        ICClogic(df, factName), file_dir['main_excel'])

                elif factName.group(0) == 'BRH':
                    new_header = pd.read_html(i.HTMLBody)[0].iloc[0]
                    df = pd.read_html(i.HTMLBody)[0][1:]
                    df.columns = new_header

                    df1 = df.drop(['LINE', 'CAP', 'Config'], axis=1)

                    new_header = ['LOB', 'HRS1', 'UPH1', 'HRS2', 'UPH2']
                    df1.columns = new_header

                    df1 = df1.fillna(0)

                    ExcelCreator.BRH1DataInsert(
                        BRHlogic(df1, factName), file_dir['main_excel'])
            else:
                print("No ICC, BRH, APCC emails workplan found")

            if Ot_EMFP != None:
                print("EMFP OT email found")
                ExcelCreator.OTDataInsert(
                    'EMFP', i.Body, file_dir['main_excel'])

            else:
                print("No EMFP OT email found")

        print("Done!")

    except PermissionError as e:
        print("Theres a problem while saving the excel file. Close file if its active.")
        print(str(e))

    except Exception as e:
        print("Error while processing data from e-mail")
        print(str(e))

    # EMFP insertion
    try:
        print("Inserting data for EMFP...")
        ExcelCreator.EMFPDataInsert(
            EMFPlogic(file_dir["EMFP"]), file_dir['main_excel'])
    except Exception as e:
        print("failed to insert EMFP data")
        print(str(e))

    # Send email to sharepoint

    # construct Outlook application instance
    olApp = win32.Dispatch('Outlook.Application')
    olNS = olApp.GetNameSpace('MAPI')

    # construct the email item object
    mailItem = olApp.CreateItem(0)
    mailItem.Subject = 'Test '  # can be any subject
    mailItem.BodyFormat = 1
    mailItem.Body = "Attachment of Consolidate View"  # can be any body
    mailItem.To = sharepoint["email"]

    # mailItem._oleobj_.Invoke(*(64209, 0, 8, 0, olNS.Accounts.Item('<email@gmail.com'))) [NOTHING JUST IGNORE FOR NOW! DONT DELETE ]

    mailItem.Attachments.Add(os.path.join(
        os.getcwd(), file_dir['main_excel']))
    # mailItem.Attachments.Add(os.path.join(os.getcwd(), r'C:\Users\Kamarul_Syamil\Desktop\Dell\Project\Test2.csv')) <*sample*>

    try:
        print("Sending excel file to sharepoint...")
        mailItem.Display()
        # mailItem.Send()

        print("Successfully sent excel file to sharepoint")

    except Exception as e:
        print(e)

    # Updated time
    wb = load_workbook(file_dir['main_excel'])
    ws = wb.active

    ws['F5'] = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    ws['F5'].alignment = Alignment(horizontal='left')
    ws['I5'] = strftime("%z", gmtime())

    wb.save(file_dir['main_excel'])


if __name__ == "__main__":
    main()
