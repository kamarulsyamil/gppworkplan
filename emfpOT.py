from email import message
import datetime
from cv2 import add
import win32com.client as client
import openpyxl

# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute
sheet = wb.active


today = datetime.date.today()

# create instance of Outlook
outlook = client.Dispatch('Outlook.Application')

# get the inbox
namespace = outlook.GetNameSpace('MAPI')
inbox = namespace.GetDefaultFolder(6)


# the email I want to download a file from

# get only mail items from the inbox (other items can exists and will return an error if you try get the subject line of a non-mail item)
mail_items = [item for item in inbox.Items if item.Class == 43]


# filter to the target email
filtered = [
    item for item in mail_items if item.Unread and item.Senton.date() == today]
# filtered1 = [item.Restrict (urn:schemas:httpmail:subject" ci_phrasematch " & "'approved'")]
# subjects=[message for message in inbox.Items if message.Subject.ci_phrasematch('EMFP Overtime')]

if len(filtered) == 0:
    print("No Attachment")
n = 0
# get the first item if it exists (assuming the there is only one item to get)
while n < len(filtered):

    if len(filtered) != 0:
        target_email = filtered[n]
        n += 1

        if target_email.Attachments.Count == 0 and n >= len(filtered):
            # for target_email in filtered:  yg ni kalau uncomment dia boleh read tapi.

            filtered1 = [
                target_email for target_email in filtered if 'EMFP Overtime' in target_email.Subject]
            p = 0
            while p < len(filtered1):
                if len(filtered1) != 0:
                    target_email = filtered1[p]
                    p += 1
                    c1 = sheet.cell(row=p, column=1)
                    # if filtered1.Count > 0:
                    print(target_email.Subject)
        # writing values to cells
                    c1.value = target_email.body
                    wb.save(
                        "C:\\Users\\Kamarul_Syamil\\Desktop\\Dell\\Project\\Test6.csv")

        elif target_email.Attachments.Count > 0:
            attachments = target_email.Attachments

    # save attachments to file
            save_path = 'C:\\Users\\Kamarul_Syamil\\Desktop\\Dell\\Project\\{}'

            for file in attachments:
                file.SaveAsFile(save_path.format(file.FileName))
                message.Unread = False

    # elif len(filtered) == 0:
    #     print ("No Email")
